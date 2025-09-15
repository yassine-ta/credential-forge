"""Excel format synthesizer using agent-generated content."""

import random
from pathlib import Path
from typing import Dict, Any

from .format_synthesizer import FormatSynthesizer
from ..utils.exceptions import SynthesizerError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelFormatSynthesizer(FormatSynthesizer):
    """Excel format synthesizer that structures agent-generated content."""
    
    def __init__(self, output_dir: str, format_type: str = 'xlsx', ultra_fast_mode: bool = False):
        """Initialize Excel format synthesizer.
        
        Args:
            output_dir: Output directory for generated files
            format_type: Excel format type (xlsx, xls, xlsm, xlsb)
            ultra_fast_mode: Enable ultra-fast mode with minimal validation
        """
        super().__init__(output_dir, ultra_fast_mode)
        self.format_type = format_type
    
    def synthesize(self, content_structure: Dict[str, Any]) -> str:
        """Structure content into Excel format.
        
        Args:
            content_structure: Generated content from agents
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Validate content structure
            self._validate_content_structure(content_structure)
            
            # Generate filename and save
            filename = self._generate_filename(content_structure)
            file_path = self._get_file_path(filename)
            
            if OPENPYXL_AVAILABLE:
                # Create Excel with openpyxl
                self._create_excel_with_openpyxl(content_structure, file_path)
            else:
                # Create simple CSV file
                self._create_simple_csv(content_structure, file_path)
            
            # Log stats
            self._log_generation_stats(content_structure)
            
            return str(file_path)
            
        except Exception as e:
            self.generation_stats['errors'] += 1
            raise SynthesizerError(f"Excel synthesis failed: {e}")
    
    def _create_excel_with_openpyxl(self, content_structure: Dict[str, Any], file_path: Path) -> None:
        """Create Excel file using openpyxl."""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets based on sections
        sections = content_structure.get('sections', [])
        credentials = content_structure.get('credentials', [])
        language = content_structure.get('language', 'en')
        
        # Title sheet
        title_sheet = wb.create_sheet("Document Info")
        self._populate_title_sheet(title_sheet, content_structure)
        
        # Data sheets
        for i, section in enumerate(sections):
            sheet_name = section.get('title', f'Sheet{i+1}')[:31]  # Excel sheet name limit
            sheet = wb.create_sheet(sheet_name)
            self._populate_data_sheet(sheet, section, language)
        
        # Credentials sheet
        if credentials:
            cred_sheet = wb.create_sheet(self._get_credentials_sheet_name(language))
            self._populate_credentials_sheet(cred_sheet, credentials, language)
        
        # Save workbook
        wb.save(str(file_path))
    
    def _populate_title_sheet(self, sheet, content_structure: Dict[str, Any]) -> None:
        """Populate the title/info sheet."""
        # Title
        sheet['A1'] = content_structure.get('title', 'Document')
        sheet['A1'].font = Font(size=16, bold=True)
        
        # Metadata
        metadata = content_structure.get('metadata', {})
        row = 3
        
        sheet[f'A{row}'] = 'Topic:'
        sheet[f'B{row}'] = metadata.get('topic', 'N/A')
        row += 1
        
        sheet[f'A{row}'] = 'Language:'
        sheet[f'B{row}'] = content_structure.get('language', 'en')
        row += 1
        
        sheet[f'A{row}'] = 'Format:'
        sheet[f'B{row}'] = content_structure.get('format_type', 'unknown')
        row += 1
        
        sheet[f'A{row}'] = 'Generated:'
        sheet[f'B{row}'] = metadata.get('generated_at', 'N/A')
        
        # Style the sheet
        for row in range(1, 7):
            sheet[f'A{row}'].font = Font(bold=True)
    
    def _populate_data_sheet(self, sheet, section: Dict[str, str], language: str) -> None:
        """Populate a data sheet with section content."""
        title = section.get('title', 'Section')
        content = section.get('content', '')
        
        # Sheet title
        sheet['A1'] = title
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Content
        lines = content.split('\n')
        for i, line in enumerate(lines, start=3):
            if line.strip():
                sheet[f'A{i}'] = line.strip()
    
    def _populate_credentials_sheet(self, sheet, credentials: list, language: str) -> None:
        """Populate the credentials sheet."""
        # Sheet title
        sheet['A1'] = self._get_credentials_sheet_name(language)
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        
        # Headers
        sheet['A3'] = self._get_credential_type_header(language)
        sheet['B3'] = self._get_credential_value_header(language)
        sheet['C3'] = self._get_credential_label_header(language)
        
        # Style headers
        for col in ['A3', 'B3', 'C3']:
            sheet[col].font = Font(bold=True)
            sheet[col].fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        
        # Credentials data
        for i, cred in enumerate(credentials, start=4):
            sheet[f'A{i}'] = cred.get('type', 'unknown')
            sheet[f'B{i}'] = cred.get('value', '')
            sheet[f'C{i}'] = cred.get('label', cred.get('type', 'unknown'))
            
            # Color code based on credential type
            if 'password' in cred.get('type', '').lower():
                sheet[f'A{i}'].fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            elif 'api' in cred.get('type', '').lower():
                sheet[f'A{i}'].fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
            else:
                sheet[f'A{i}'].fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    
    def _get_credentials_sheet_name(self, language: str) -> str:
        """Get localized credentials sheet name."""
        names = {
            'en': 'Credentials',
            'fr': 'Identifiants',
            'es': 'Credenciales',
            'de': 'Anmeldedaten',
            'it': 'Credenziali'
        }
        return names.get(language, names['en'])
    
    def _get_credential_type_header(self, language: str) -> str:
        """Get localized credential type header."""
        headers = {
            'en': 'Type',
            'fr': 'Type',
            'es': 'Tipo',
            'de': 'Typ',
            'it': 'Tipo'
        }
        return headers.get(language, headers['en'])
    
    def _get_credential_value_header(self, language: str) -> str:
        """Get localized credential value header."""
        headers = {
            'en': 'Value',
            'fr': 'Valeur',
            'es': 'Valor',
            'de': 'Wert',
            'it': 'Valore'
        }
        return headers.get(language, headers['en'])
    
    def _get_credential_label_header(self, language: str) -> str:
        """Get localized credential label header."""
        headers = {
            'en': 'Label',
            'fr': 'Étiquette',
            'es': 'Etiqueta',
            'de': 'Bezeichnung',
            'it': 'Etichetta'
        }
        return headers.get(language, headers['en'])
    
    def _create_simple_csv(self, content_structure: Dict[str, Any], file_path: Path) -> None:
        """Create simple CSV file."""
        import csv
        
        with open(file_path.with_suffix('.csv'), 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Title
            writer.writerow([content_structure.get('title', 'Document')])
            writer.writerow([])
            
            # Sections
            sections = content_structure.get('sections', [])
            for section in sections:
                writer.writerow([section.get('title', 'Section')])
                content = section.get('content', '')
                for line in content.split('\n'):
                    if line.strip():
                        writer.writerow([line.strip()])
                writer.writerow([])
            
            # Credentials
            credentials = content_structure.get('credentials', [])
            if credentials:
                writer.writerow(['Credentials'])
                for cred in credentials:
                    writer.writerow([cred.get('type', ''), cred.get('value', ''), cred.get('label', '')])
    
    def _generate_filename(self, content_structure: Dict[str, Any]) -> str:
        """Generate Excel filename."""
        title = content_structure.get('title', 'spreadsheet')
        timestamp = self._get_current_timestamp()
        random_id = random.randint(1000, 9999)
        
        # Clean title for filename
        clean_title = "".join(c for c in title if c.isalnum() or c in (' ', '-', '_')).rstrip()
        clean_title = clean_title.replace(' ', '_').lower()
        
        return f"spreadsheet_{clean_title}_{timestamp}_{random_id}.{self.format_type}"
    
    def _get_current_timestamp(self) -> str:
        """Get current timestamp for filename."""
        from datetime import datetime
        return datetime.now().strftime('%Y%m%d_%H%M%S')
